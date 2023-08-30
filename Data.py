import openpyxl 

class Data:

    def __init__(self, path:str) -> None:
        self.excel_document = openpyxl.load_workbook(path)

    def get_address(self)->list:
        
        self.sheet = self.excel_document.get_sheet_by_name('chihuahua')

        address = []

        for column in self.sheet.iter_cols():
            column_name = column[0].value
            if column_name == 'direccion':
                for cell in column:
                    if cell.value != 'direccion':
                        address.append(cell.value)

        return address

    def add_address_to_excel(self, list_address:list):

        try:

            for i, address in enumerate(list_address):

                self.sheet.cell(column=15, row= i+1, value= address)

                self.excel_document.save('./Muestra Att Nickmare 29082023.xlsx')
                
        except Exception as e:
            print(e)

    
if __name__ == '__main__':
    data = Data('./Chihuahua1.xlsx')
    print(data.get_address())
    # data.add_values()